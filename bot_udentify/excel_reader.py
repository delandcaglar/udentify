import os
import pandas as pd
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill
from openpyxl.styles import colors
from openpyxl.styles import  Color,PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook #for not overwriting
import csv
#Not Mağza ve tarıh kısmını eklemen gerek
##hata olursa satir 55 e bak  neden commend out yaptin panda dataframe i

#folder_count = 0  # type: int ##-1 cunku macleerde gizli folder var




#Avangate_ID olayini arastie


print('Variables____________')
firm_path = "/Users/ilkedelandcaglar/Downloads/udentify/bot_udentify/ıstenen_magazalar.xlsx"
excel_file_1 = '/Users/ilkedelandcaglar/Downloads/udentify/bot_udentify/ıstenen_magazalar.xlsx'
alphabet_list=["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R"] #18 tane var
ornek_liste = ['9.9-Oto Koltukları', 6.87, 393, 7447, 95, 66.81, 154, 35.38, -16.93, 1.56, 60, 36, -40, 79449.76, 53735.16, -32.37, 1.15, 1.92]#6
ornek_liste1 = ['10.9-Eğitici (oto koltuğu yanı)', 3.5, 28, 6403, 32, 13.37, -2, 30.13, -16.23, 27.62, 455, 466, 2.42, 14549.77, 14282.51, -1.84, 0.59, 1.18]#7
ornek_liste2 = ['5.4-Yenidoğan Duvar', 2.04, 195, 7187, 130, 10.32, 29, 16.7, 2.05, '-', '-', '-', '-', '-', '-', '-', '-', 0.64]#8
ornek_liste3 = ['Kasa', 8.68, 460, 14941, 147, 25.8, 127, 27.85, 2.85, '-', '-', '-', '-', '-', '-', '-', '-', 3.57]#9
ornek_buyuk_liste = [ornek_liste,ornek_liste1,ornek_liste2,ornek_liste3]
folder_list = []
print('Variables_________')


print('Functions__________________')


print('Functions__________________')




def update_excel_file(excel_file_1,istenilen_adet):




    excel_file_1 = excel_file_1
    #df_first_shift = pd.read_excel(excel_file_1, sheet_name='Sheet')

    #workbook = Workbook() #overvrites

    workbook = load_workbook ( excel_file_1 ) #JUST SAVES
    ws = workbook.active


    sayi_katsayi = 0
    while sayi_katsayi < istenilen_adet:
        sabit_katsayi = ( 4 * sayi_katsayi)

        ws[f"B{1+sabit_katsayi}"] = str('Adı')
        ws[f"B{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"C{1+sabit_katsayi}"] = str('Ürün kodu')
        ws[f"C{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"D{1+sabit_katsayi}"] = str('Fiyat($)')
        ws[f"D{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"E{1+sabit_katsayi}"] = str('Deri Rengi')
        ws[f"E{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"F{1+sabit_katsayi}"] = str('Deri Rengi Opsiyonları')
        ws[f"F{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"G{1+sabit_katsayi}"] = str('Deri Türü')
        ws[f"G{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"H{1+sabit_katsayi}"] = str('Deri Oranı')
        ws[f"H{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"I{1+sabit_katsayi}"] = str('Tüy Rengi')
        ws[f"I{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"J{1+sabit_katsayi}"] = str('Tüy Rengi Opsiyonları')
        ws[f"J{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"K{1+sabit_katsayi}"] = str('Sezon')
        ws[f"K{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"L{1+sabit_katsayi}"] = str('Bakım')
        ws[f"L{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"M{1+sabit_katsayi}"] = str('Fermuar Türü')
        ws[f"M{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"N{1+sabit_katsayi}"] = str('Fermuar Materyeli')
        ws[f"N{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"O{1+sabit_katsayi}"] = str('Düğmeleri')
        ws[f"O{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"P{1+sabit_katsayi}"] = str('Çift Taraflı')
        ws[f"P{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"Q{1+sabit_katsayi}"] = str('Koleksiyon Adı')
        ws[f"Q{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"R{1+sabit_katsayi}"] = str('Cep Türü')
        ws[f"R{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"S{1+sabit_katsayi}"] = str('İsim Etiketi')
        ws[f"S{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"T{1+sabit_katsayi}"] = str('Ağırlık')
        ws[f"T{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"U{1+sabit_katsayi}"] = str('Açıklama')
        ws[f"U{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"V{1+sabit_katsayi}"] = str('Materyaller')
        ws[f"V{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"W{1+sabit_katsayi}"] = str('Ürün Ölçüleri(cm)')
        ws[f"W{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"X{1+sabit_katsayi}"] = str('İç Kısım')
        ws[f"X{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"Y{1+sabit_katsayi}"] = str('Dış Kısım')
        ws[f"Y{1+sabit_katsayi}"].font = Font(bold=True)
        ws[f"Z{1+sabit_katsayi}"] = str('Jeket Türü')
        ws[f"Z{1+sabit_katsayi}"].font = Font(bold=True)


        # ingilizce kisim
        ws[f"B{3 + sabit_katsayi}"] = str ( 'Name' )
        ws[f"B{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"C{3 + sabit_katsayi}"] = str ( 'Product Code' )
        ws[f"C{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"D{3 + sabit_katsayi}"] = str ( 'Price($)' )
        ws[f"D{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"E{3 + sabit_katsayi}"] = str ( 'Colour of The Leather' )
        ws[f"E{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"F{3 + sabit_katsayi}"] = str ( 'Colour Options Of Leather' )
        ws[f"F{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"G{3 + sabit_katsayi}"] = str ( 'Type Of Leather' )
        ws[f"G{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"H{3 + sabit_katsayi}"] = str ( 'Leather Percentage' )
        ws[f"H{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"I{3 + sabit_katsayi}"] = str ( 'Feather Color' )
        ws[f"I{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"J{3 + sabit_katsayi}"] = str ( 'Feather Color Options' )
        ws[f"J{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"K{3 + sabit_katsayi}"] = str ( 'Seeason' )
        ws[f"K{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"L{3 + sabit_katsayi}"] = str ( 'Care' )
        ws[f"L{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"M{3 + sabit_katsayi}"] = str ( 'Zipper Type' )
        ws[f"M{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"N{3 + sabit_katsayi}"] = str ( 'Zipper Material' )
        ws[f"N{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"O{3 + sabit_katsayi}"] = str ( 'Buttons' )
        ws[f"O{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"P{3 + sabit_katsayi}"] = str ( 'Reversable' )
        ws[f"P{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"Q{3 + sabit_katsayi}"] = str ( 'Collection Name' )
        ws[f"Q{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"R{3 + sabit_katsayi}"] = str ( 'Pocket Type' )
        ws[f"R{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"S{3 + sabit_katsayi}"] = str ( 'Name Tag' )
        ws[f"S{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"T{3 + sabit_katsayi}"] = str ( 'Weight' )
        ws[f"T{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"U{3 + sabit_katsayi}"] = str ( 'Description' )
        ws[f"U{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"V{3 + sabit_katsayi}"] = str ( 'Materials' )
        ws[f"V{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"W{3 + sabit_katsayi}"] = str ( 'Product Meeasurements' )
        ws[f"W{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"X{3 + sabit_katsayi}"] = str ( 'Interior' )
        ws[f"X{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"Y{3 + sabit_katsayi}"] = str ( 'Exterior' )
        ws[f"Y{3 + sabit_katsayi}"].font = Font ( bold=True )
        ws[f"Z{3 + sabit_katsayi}"] = str ( 'Type of Jacket' )
        ws[f"Z{3 + sabit_katsayi}"].font = Font ( bold=True )

        sayi_katsayi += 1



    ##Adjustment for excel files
    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=25)

    ws.column_dimensions = dim_holder


    #
    workbook.save(filename = excel_file_1)#hatali workbbook nerden cikti


def tum_dosyalara_eris(firm_path,ornek_buyuk_liste): ##!!!!!!!!!! apple disi sistemlerde sikinti cikarabilir
    folder_list_counter = 1
    loop_starteer = 0
    folder_count = 0
    input_path = firm_path  # type: str
    for folders in os.listdir(input_path):
        folder_count += 1  # increment counter
        folder_list.append(folders)
        print(folders) #ds.store yuzunden -1 yapacaksin  ##buradan sonrasini yaz

    print("There are {0} folders".format(folder_count-1))# ds.store icin eksi 1 koymak zorundasin
    real_folder_numer = folder_count-1
    print(folder_list)# birincisi her zaman DS_Store
    print(folder_list[1])

    ##sql data cekme fonksiyonunu buraya entegre edip  bbutuk listeleri eslestireceksin !!!!!!!!!!!!!!!!! ve biteceek

    while loop_starteer < real_folder_numer:
        dynamic_variable = f'/Users/ilkedelandcaglar/Downloads/udentıfy/Firms/{folder_list[folder_list_counter]}/{folder_list[folder_list_counter]}Excel.xlsx'
        print(dynamic_variable)
        update_excel_file ( dynamic_variable, ornek_buyuk_liste )

        folder_list_counter = folder_list_counter +1
        loop_starteer = loop_starteer+1

#tum_dosyalara_eris(firm_path,ornek_buyuk_liste)
#dynamic_variable = f'/Users/ilkedelandcaglar/Downloads/udentıfy/Firms/{Ebebek}/{Ebebek}Excel.xlsx'


def variable_excel_file(excel_file_1,istenilen_adet):
    final_listesi = []


    workbook = load_workbook ( excel_file_1 )  # JUST SAVES
    ws = workbook.active


    sayi_katsayi = 0
    while sayi_katsayi < istenilen_adet:
        sabit_katsayi = ( 1 * sayi_katsayi)




        Firma = ws[f"A{2+sabit_katsayi}"].internal_value #('Açıklama')
        print(Firma)
        Magaza_Adı = ws[f"B{2 + sabit_katsayi}"].internal_value  # ('Açıklama')
        print ( Magaza_Adı )
        Magaza_no = ws[f"C{2 + sabit_katsayi}"].internal_value # ('Açıklama')
        print ( Magaza_no )
        ilk_tarih = ws[f"D{2 + sabit_katsayi}"].internal_value  # ('Açıklama')
        print ( ilk_tarih )
        son_tarih = ws[f"E{2 + sabit_katsayi}"].internal_value  # ('Açıklama')
        print ( son_tarih )

        magaza_id_yo = ws[f"F{2 + sabit_katsayi}"].internal_value  # ('Açıklama')
        print ( magaza_id_yo )
        isim_listesi = ws[f"G{2 + sabit_katsayi}"].internal_value  # ('Açıklama')
        print ( isim_listesi )





        bos_liste = [Firma,Magaza_Adı,Magaza_no,(ilk_tarih),son_tarih,magaza_id_yo,isim_listesi]
        print(bos_liste)
        print ( bos_liste[3] )
        #final_listesi.append()


        sayi_katsayi += 1





istenilen_adet = 3

# yaz
#update_excel_file( firm_path,istenilen_adet )

variable_excel_file(firm_path,istenilen_adet)

#oku

















# import csv
# fields=['','first','second','third']
# with open(csv_dosyasi, 'a') as f:
#     writer = csv.writer(f)
#     writer.writerow(fields)



















#TESTING CODES____________________________________________________________________________
#FOR GERTTTING THEE COLOMB DATA
'''excel_str = "/Users/ilkedelandcaglar/Downloads/udentıfy/Firms/Ebebek/EbebekExcel.xlsx"

##This is for testing purposes
df = pd.read_excel(excel_str, sheet_name='Data')
print(df.iloc[7]) #+2 koyular
elma = (df.iloc[7])
print(elma.tolist())'''

##FOR GETTING THE ROW DATA
'''def excel_manipulate(excel_str):
    email_list = pd.read_excel(excel_str)
    all_names = email_list['Mağaza']
    print(all_names)
    print(len(all_names))

#excel_manipulate(excel_str)'''